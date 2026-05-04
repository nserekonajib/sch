# exams.py - Fixed with only exam_date (no start_date/end_date)
from flask import Blueprint, render_template, request, jsonify, session, send_file
from supabase import create_client, Client
import os
import uuid
import random
import string
from datetime import datetime, timedelta
import json
import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from functools import wraps
from dotenv import load_dotenv

load_dotenv()

# Initialize Supabase client
SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_KEY')
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

exams_bp = Blueprint('exams', __name__, url_prefix='/exams')

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            return jsonify({'success': False, 'message': 'Please login'}), 401
        return f(*args, **kwargs)
    return decorated_function

def get_institute(user_id):
    try:
        response = supabase.table('institutes')\
            .select('*')\
            .eq('user_id', user_id)\
            .execute()
        
        if response.data and len(response.data) > 0:
            return response.data[0]
        return None
    except Exception as e:
        print(f"Error getting institute: {e}")
        return None

@exams_bp.route('/')
@login_required
def index():
    user = session.get('user')
    institute = get_institute(user['id'])
    
    if not institute:
        return render_template('exams/index.html', exams=[], institute=None, datetime=datetime)
    
    try:
        # Get all exams with optional date filtering
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        query = supabase.table('exams')\
            .select('*')\
            .eq('institute_id', institute['id'])
        
        if start_date:
            query = query.gte('exam_date', start_date)
        if end_date:
            query = query.lte('exam_date', end_date)
        
        exams_response = query.order('exam_date', desc=True).execute()
        
        exams = exams_response.data if exams_response.data else []
        
        return render_template('exams/index.html', exams=exams, institute=institute, datetime=datetime)
        
    except Exception as e:
        print(f"Error loading exams page: {e}")
        return render_template('exams/index.html', exams=[], institute=institute, datetime=datetime)

@exams_bp.route('/marks')
@login_required
def marks():
    user = session.get('user')
    institute = get_institute(user['id'])
    
    if not institute:
        return render_template('exams/marks.html', exams=[], classes=[], institute=None, current_year=datetime.now().year)
    
    try:
        # Get all exams
        exams_response = supabase.table('exams')\
            .select('*')\
            .eq('institute_id', institute['id'])\
            .order('exam_date', desc=True)\
            .execute()
        
        exams = exams_response.data if exams_response.data else []
        
        # Get all classes
        classes_response = supabase.table('classes')\
            .select('*')\
            .eq('institute_id', institute['id'])\
            .order('name')\
            .execute()
        
        classes = classes_response.data if classes_response.data else []
        
        return render_template('exams/marks.html', exams=exams, classes=classes, institute=institute, current_year=datetime.now().year)
        
    except Exception as e:
        print(f"Error loading marks page: {e}")
        return render_template('exams/marks.html', exams=[], classes=[], institute=institute, current_year=datetime.now().year)

@exams_bp.route('/api/exams', methods=['GET'])
@login_required
def get_exams():
    user = session.get('user')
    institute = get_institute(user['id'])
    
    if not institute:
        return jsonify({'success': False, 'message': 'Institute not found'}), 400
    
    try:
        # Get all exams with optional date filtering
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        query = supabase.table('exams')\
            .select('*')\
            .eq('institute_id', institute['id'])
        
        if start_date:
            query = query.gte('exam_date', start_date)
        if end_date:
            query = query.lte('exam_date', end_date)
        
        response = query.order('exam_date', desc=True).execute()
        
        exams = response.data if response.data else []
        
        return jsonify({'success': True, 'exams': exams})
        
    except Exception as e:
        print(f"Error getting exams: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@exams_bp.route('/api/exams/create', methods=['POST'])
@login_required
def create_exam():
    user = session.get('user')
    institute = get_institute(user['id'])
    
    if not institute:
        return jsonify({'success': False, 'message': 'Institute not found'}), 400
    
    try:
        data = request.get_json()
        exam_name = data.get('exam_name', '').strip()
        total_marks = float(data.get('total_marks', 0))
        exam_date = data.get('exam_date')
        
        if not exam_name:
            return jsonify({'success': False, 'message': 'Exam name is required'}), 400
        if total_marks <= 0:
            return jsonify({'success': False, 'message': 'Total marks must be greater than 0'}), 400
        if not exam_date:
            return jsonify({'success': False, 'message': 'Exam date is required'}), 400
        
        exam_id = str(uuid.uuid4())
        exam_data = {
            'id': exam_id,
            'institute_id': institute['id'],
            'exam_name': exam_name,
            'total_marks': total_marks,
            'exam_date': exam_date,
            'is_published': False,
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat()
        }
        
        result = supabase.table('exams').insert(exam_data).execute()
        
        if result.data:
            return jsonify({'success': True, 'message': 'Exam created successfully', 'exam': result.data[0]})
        else:
            return jsonify({'success': False, 'message': 'Failed to create exam'}), 500
            
    except Exception as e:
        print(f"Error creating exam: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@exams_bp.route('/api/exams/date-range', methods=['GET'])
@login_required
def get_exam_date_range():
    """Get min and max exam dates for filtering"""
    user = session.get('user')
    institute = get_institute(user['id'])
    
    if not institute:
        return jsonify({'success': False, 'message': 'Institute not found'}), 400
    
    try:
        response = supabase.table('exams')\
            .select('exam_date')\
            .eq('institute_id', institute['id'])\
            .execute()
        
        dates = [e['exam_date'] for e in response.data if e.get('exam_date')] if response.data else []
        
        if dates:
            min_date = min(dates)
            max_date = max(dates)
        else:
            min_date = datetime.now().date().isoformat()
            max_date = datetime.now().date().isoformat()
        
        return jsonify({'success': True, 'min_date': min_date, 'max_date': max_date})
        
    except Exception as e:
        print(f"Error getting date range: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@exams_bp.route('/api/exams/<exam_id>/toggle-publish', methods=['PUT'])
@login_required
def toggle_publish(exam_id):
    user = session.get('user')
    institute = get_institute(user['id'])
    
    if not institute:
        return jsonify({'success': False, 'message': 'Institute not found'}), 400
    
    try:
        data = request.get_json()
        is_published = data.get('is_published', False)
        
        result = supabase.table('exams')\
            .update({'is_published': is_published, 'updated_at': datetime.now().isoformat()})\
            .eq('id', exam_id)\
            .eq('institute_id', institute['id'])\
            .execute()
        
        if result.data:
            return jsonify({'success': True, 'message': f'Exam {"published" if is_published else "unpublished"} successfully'})
        else:
            return jsonify({'success': False, 'message': 'Exam not found'}), 404
            
    except Exception as e:
        print(f"Error toggling publish: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@exams_bp.route('/api/exams/<exam_id>', methods=['DELETE'])
@login_required
def delete_exam(exam_id):
    user = session.get('user')
    institute = get_institute(user['id'])
    
    if not institute:
        return jsonify({'success': False, 'message': 'Institute not found'}), 400
    
    try:
        supabase.table('exam_marks')\
            .delete()\
            .eq('exam_id', exam_id)\
            .eq('institute_id', institute['id'])\
            .execute()
        
        result = supabase.table('exams')\
            .delete()\
            .eq('id', exam_id)\
            .eq('institute_id', institute['id'])\
            .execute()
        
        if result.data:
            return jsonify({'success': True, 'message': 'Exam deleted successfully'})
        else:
            return jsonify({'success': False, 'message': 'Exam not found'}), 404
            
    except Exception as e:
        print(f"Error deleting exam: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500



@exams_bp.route('/api/subjects', methods=['GET'])
@login_required
def get_subjects():
    user = session.get('user')
    institute = get_institute(user['id'])
    
    if not institute:
        return jsonify({'success': False, 'message': 'Institute not found'}), 400
    
    try:
        class_id = request.args.get('class_id')
        
        if not class_id:
            return jsonify({'success': False, 'message': 'Class ID required'}), 400
        
        response = supabase.table('class_subjects')\
            .select('*, subjects(name)')\
            .eq('class_id', class_id)\
            .eq('institute_id', institute['id'])\
            .execute()
        
        print(f"Subjects query result for class {class_id}: {response.data}")  # Debug log
        
        subjects = response.data if response.data else []
        
        # Format subjects for frontend
        formatted_subjects = []
        for subject in subjects:
            formatted_subjects.append({
                'id': subject['subject_id'],
                'name': subject['subjects']['name'] if subject.get('subjects') else 'Unknown',
                'marks': subject['marks'],
                'teacher_id': subject.get('teacher_id')
            })
        
        return jsonify({'success': True, 'subjects': formatted_subjects})
        
    except Exception as e:
        print(f"Error getting subjects: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    
# @exams_bp.route('/api/debug/class-data', methods=['GET'])
# @login_required
# def debug_class_data():
#     """Debug endpoint to check class data"""
#     user = session.get('user')
#     institute = get_institute(user['id'])
    
#     if not institute:
#         return jsonify({'success': False, 'message': 'Institute not found'}), 400
    
#     try:
#         class_id = request.args.get('class_id')
        
#         if not class_id:
#             return jsonify({'success': False, 'message': 'Class ID required'}), 400
        
#         debug_info = {}
        
#         # 1. Check class exists
#         class_response = supabase.table('classes')\
#             .select('*')\
#             .eq('id', class_id)\
#             .execute()
#         debug_info['class'] = class_response.data[0] if class_response.data else None
        
#         # 2. Check class_subjects
#         subjects_response = supabase.table('class_subjects')\
#             .select('*, subjects(name)')\
#             .eq('class_id', class_id)\
#             .eq('institute_id', institute['id'])\
#             .execute()
#         debug_info['class_subjects'] = subjects_response.data if subjects_response.data else []
#         debug_info['class_subjects_count'] = len(subjects_response.data) if subjects_response.data else 0
        
#         # 3. Check if subjects table has names
#         subject_ids = [s['subject_id'] for s in debug_info['class_subjects']] if debug_info['class_subjects'] else []
#         if subject_ids:
#             subjects_table = supabase.table('subjects')\
#                 .select('id, name')\
#                 .in_('id', subject_ids)\
#                 .execute()
#             debug_info['subjects_table'] = subjects_table.data if subjects_table.data else []
        
#         # 4. Check class enrollments
#         enrollments_response = supabase.table('class_enrollments')\
#             .select('student_id, academic_year')\
#             .eq('class_id', class_id)\
#             .eq('academic_year', 2026)\
#             .execute()
#         debug_info['enrollments'] = enrollments_response.data if enrollments_response.data else []
#         debug_info['enrollments_count'] = len(enrollments_response.data) if enrollments_response.data else 0
        
#         # 5. Get student details if any
#         if debug_info['enrollments']:
#             student_ids = [e['student_id'] for e in debug_info['enrollments']]
#             students_response = supabase.table('students')\
#                 .select('id, name, student_id')\
#                 .eq('institute_id', institute['id'])\
#                 .in_('id', student_ids)\
#                 .execute()
#             debug_info['students'] = students_response.data if students_response.data else []
        
#         return jsonify({'success': True, 'debug': debug_info})
        
#     except Exception as e:
#         print(f"Debug error: {e}")
#         import traceback
#         traceback.print_exc()
#         return jsonify({'success': False, 'message': str(e), 'traceback': traceback.format_exc()}), 500
    
    
@exams_bp.route('/api/class-students', methods=['GET'])
@login_required
def get_class_students():
    user = session.get('user')
    institute = get_institute(user['id'])
    
    if not institute:
        return jsonify({'success': False, 'message': 'Institute not found'}), 400
    
    try:
        class_id = request.args.get('class_id')
        academic_year = request.args.get('academic_year', str(datetime.now().year))
        
        if not class_id:
            return jsonify({'success': False, 'message': 'Class ID required'}), 400
        
        enrollments_response = supabase.table('class_enrollments')\
            .select('student_id')\
            .eq('class_id', class_id)\
            .eq('academic_year', int(academic_year))\
            .execute()
        
        student_ids = [e['student_id'] for e in enrollments_response.data] if enrollments_response.data else []
        
        if not student_ids:
            return jsonify({'success': True, 'students': []})
        
        students_response = supabase.table('students')\
            .select('id, name, student_id')\
            .eq('institute_id', institute['id'])\
            .in_('id', student_ids)\
            .order('name')\
            .execute()
        
        students = students_response.data if students_response.data else []
        
        return jsonify({'success': True, 'students': students})
        
    except Exception as e:
        print(f"Error getting class students: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    
@exams_bp.route('/api/marks', methods=['GET'])
@login_required
def get_marks():
    """Get marks for an exam and class for a specific academic year"""
    user = session.get('user')
    institute = get_institute(user['id'])
    
    if not institute:
        return jsonify({'success': False, 'message': 'Institute not found'}), 400
    
    try:
        exam_id = request.args.get('exam_id')
        class_id = request.args.get('class_id')
        academic_year = request.args.get('academic_year', str(datetime.now().year))
        version = request.args.get('version', 'current')
        history_date = request.args.get('history_date')
        
        if not exam_id or not class_id:
            return jsonify({'success': False, 'message': 'Exam ID and Class ID required'}), 400
        
        # Get exam details for total marks
        exam_response = supabase.table('exams')\
            .select('total_marks, exam_name')\
            .eq('id', exam_id)\
            .execute()
        
        if not exam_response.data:
            return jsonify({'success': False, 'message': 'Exam not found'}), 404
        
        exam = exam_response.data[0]
        exam_total_marks = exam['total_marks']
        
        # Get subjects for the class
        subjects_response = supabase.table('class_subjects')\
            .select('*, subjects!inner(name)')\
            .eq('class_id', class_id)\
            .eq('institute_id', institute['id'])\
            .execute()
        
        subjects = subjects_response.data if subjects_response.data else []
        
        # Format subjects for frontend
        formatted_subjects = []
        for subject in subjects:
            formatted_subjects.append({
                'id': subject['subject_id'],
                'name': subject['subjects']['name'] if subject.get('subjects') else 'Unknown',
                'max_marks': subject['marks']
            })
        
        # If no subjects found, return early
        if not formatted_subjects:
            return jsonify({
                'success': True,
                'students': [],
                'subjects': [],
                'exam_total_marks': exam_total_marks,
                'message': 'No subjects assigned to this class'
            })
        
        # Get students enrolled in this specific class for the academic year
        enrollments_response = supabase.table('class_enrollments')\
            .select('student_id')\
            .eq('class_id', class_id)\
            .eq('academic_year', int(academic_year))\
            .execute()
        
        student_ids = [e['student_id'] for e in enrollments_response.data] if enrollments_response.data else []
        
        # IMPORTANT: If no students found, show ALL students with a warning
        # Don't auto-enroll because of unique constraint!
        if not student_ids:
            print(f"No students enrolled in class {class_id} for {academic_year}")
            
            # Get ALL students from the institute
            all_students_response = supabase.table('students')\
                .select('id, name, student_id')\
                .eq('institute_id', institute['id'])\
                .order('name')\
                .execute()
            
            all_students = all_students_response.data if all_students_response.data else []
            
            # Return all students but mark them as not enrolled
            marks_data = []
            for student in all_students:
                student_marks = {
                    'student_id': student['id'],
                    'student_name': student['name'],
                    'student_number': student['student_id'],
                    'subjects': [],
                    'enrollment_status': 'not_enrolled',  # Add status
                    'total_obtained': 0,
                    'exam_total_marks': exam_total_marks,
                    'percentage': 0
                }
                
                # Add empty marks for all subjects
                for subject in formatted_subjects:
                    student_marks['subjects'].append({
                        'subject_id': subject['id'],
                        'subject_name': subject['name'],
                        'max_marks': subject['max_marks'],
                        'obtained': None
                    })
                
                marks_data.append(student_marks)
            
            return jsonify({
                'success': True,
                'students': marks_data,
                'subjects': formatted_subjects,
                'exam_total_marks': exam_total_marks,
                'warning': f'No students enrolled in this class for {academic_year}. Showing all students.'
            })
        
        # Get student details for enrolled students
        students_response = supabase.table('students')\
            .select('id, name, student_id')\
            .eq('institute_id', institute['id'])\
            .in_('id', student_ids)\
            .order('name')\
            .execute()
        
        students = students_response.data if students_response.data else []
        
        # Get marks based on version
        if version == 'historical' and history_date:
            marks_response = supabase.table('exam_marks_history')\
                .select('*')\
                .eq('exam_id', exam_id)\
                .eq('class_id', class_id)\
                .eq('institute_id', institute['id'])\
                .eq('record_date', history_date)\
                .in_('student_id', student_ids)\
                .execute()
        else:
            if student_ids:
                marks_response = supabase.table('exam_marks')\
                    .select('*')\
                    .eq('exam_id', exam_id)\
                    .eq('class_id', class_id)\
                    .eq('institute_id', institute['id'])\
                    .in_('student_id', student_ids)\
                    .execute()
            else:
                marks_response = supabase.table('exam_marks')\
                    .select('*')\
                    .eq('exam_id', exam_id)\
                    .eq('class_id', class_id)\
                    .eq('institute_id', institute['id'])\
                    .execute()
        
        existing_marks = {}
        for mark in marks_response.data if marks_response.data else []:
            key = f"{mark['student_id']}_{mark['subject_id']}"
            existing_marks[key] = mark
        
        # Prepare marks data
        marks_data = []
        for student in students:
            student_marks = {
                'student_id': student['id'],
                'student_name': student['name'],
                'student_number': student['student_id'],
                'subjects': [],
                'enrollment_status': 'enrolled'
            }
            
            # Calculate total obtained marks across all subjects
            total_obtained = 0
            
            for subject in formatted_subjects:
                subject_id = subject['id']
                subject_name = subject['name']
                max_marks = subject['max_marks']
                key = f"{student['id']}_{subject_id}"
                
                if key in existing_marks:
                    obtained = existing_marks[key]['obtained_marks']
                else:
                    obtained = None
                
                student_marks['subjects'].append({
                    'subject_id': subject_id,
                    'subject_name': subject_name,
                    'max_marks': max_marks,
                    'obtained': obtained
                })
                
                if obtained:
                    total_obtained += obtained
            
            # Calculate percentage based on EXAM TOTAL MARKS
            percentage = round((total_obtained / exam_total_marks * 100), 1) if exam_total_marks > 0 else 0
            
            student_marks['total_obtained'] = total_obtained
            student_marks['exam_total_marks'] = exam_total_marks
            student_marks['percentage'] = percentage
            marks_data.append(student_marks)
        
        return jsonify({
            'success': True,
            'students': marks_data,
            'subjects': formatted_subjects,
            'exam_total_marks': exam_total_marks
        })
        
    except Exception as e:
        print(f"Error getting marks: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500
    
    
@exams_bp.route('/api/marks/save', methods=['POST'])
@login_required
def save_marks():
    user = session.get('user')
    institute = get_institute(user['id'])
    
    if not institute:
        return jsonify({'success': False, 'message': 'Institute not found'}), 400
    
    try:
        data = request.get_json()
        exam_id = data.get('exam_id')
        class_id = data.get('class_id')
        academic_year = data.get('academic_year', datetime.now().year)
        marks_data = data.get('marks', [])
        
        if not exam_id or not class_id:
            return jsonify({'success': False, 'message': 'Exam ID and Class ID required'}), 400
        
        saved_count = 0
        errors = []
        
        for mark_entry in marks_data:
            try:
                student_id = mark_entry.get('student_id')
                subject_id = mark_entry.get('subject_id')
                obtained_marks = mark_entry.get('obtained_marks')
                
                if obtained_marks is None or obtained_marks == '':
                    continue
                
                obtained_marks = float(obtained_marks)
                
                # Check if record exists
                existing = supabase.table('exam_marks')\
                    .select('*')\
                    .eq('exam_id', exam_id)\
                    .eq('class_id', class_id)\
                    .eq('student_id', student_id)\
                    .eq('subject_id', subject_id)\
                    .eq('institute_id', institute['id'])\
                    .execute()
                
                if existing.data:
                    # Save to history before updating
                    old_mark = supabase.table('exam_marks')\
                        .select('*')\
                        .eq('id', existing.data[0]['id'])\
                        .execute()
                    
                    if old_mark.data:
                        history_data = {
                            'id': str(uuid.uuid4()),
                            'institute_id': institute['id'],
                            'exam_id': exam_id,
                            'class_id': class_id,
                            'student_id': student_id,
                            'subject_id': subject_id,
                            'obtained_marks': old_mark.data[0]['obtained_marks'],
                            'record_date': datetime.now().date().isoformat(),
                            'created_at': datetime.now().isoformat()
                        }
                        supabase.table('exam_marks_history').insert(history_data).execute()
                    
                    # Update existing
                    supabase.table('exam_marks')\
                        .update({
                            'obtained_marks': obtained_marks,
                            'updated_at': datetime.now().isoformat()
                        })\
                        .eq('id', existing.data[0]['id'])\
                        .execute()
                else:
                    # Insert new - also save initial entry to history
                    mark_id = str(uuid.uuid4())
                    mark_data = {
                        'id': mark_id,
                        'institute_id': institute['id'],
                        'exam_id': exam_id,
                        'class_id': class_id,
                        'student_id': student_id,
                        'subject_id': subject_id,
                        'obtained_marks': obtained_marks,
                        'created_at': datetime.now().isoformat(),
                        'updated_at': datetime.now().isoformat()
                    }
                    supabase.table('exam_marks').insert(mark_data).execute()
                    
                    # Save initial entry to history as well
                    history_data = {
                        'id': str(uuid.uuid4()),
                        'institute_id': institute['id'],
                        'exam_id': exam_id,
                        'class_id': class_id,
                        'student_id': student_id,
                        'subject_id': subject_id,
                        'obtained_marks': obtained_marks,
                        'record_date': datetime.now().date().isoformat(),
                        'created_at': datetime.now().isoformat()
                    }
                    supabase.table('exam_marks_history').insert(history_data).execute()
                
                saved_count += 1
                
            except Exception as e:
                errors.append(f"Error saving marks for student {mark_entry.get('student_id')}: {str(e)}")
        
        return jsonify({
            'success': True,
            'message': f'Saved {saved_count} mark(s) successfully',
            'errors': errors if errors else None
        })
        
    except Exception as e:
        print(f"Error saving marks: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
# Fix the get_history_dates endpoint in exams.py

@exams_bp.route('/api/marks/history/dates', methods=['GET'])
@login_required
def get_history_dates():
    """Get available history dates for an exam and class"""
    user = session.get('user')
    institute = get_institute(user['id'])
    
    if not institute:
        return jsonify({'success': False, 'message': 'Institute not found'}), 400
    
    try:
        exam_id = request.args.get('exam_id')
        class_id = request.args.get('class_id')
        
        print(f"Getting history dates for exam: {exam_id}, class: {class_id}")
        
        if not exam_id or not class_id:
            return jsonify({'success': False, 'message': 'Exam ID and Class ID required'}), 400
        
        # Query the history table for distinct record dates
        response = supabase.table('exam_marks_history')\
            .select('record_date')\
            .eq('exam_id', exam_id)\
            .eq('class_id', class_id)\
            .eq('institute_id', institute['id'])\
            .execute()
        
        print(f"History query response: {response.data}")
        
        # Get unique dates and sort them
        dates_set = set()
        for record in response.data if response.data else []:
            if record.get('record_date'):
                dates_set.add(record['record_date'])
        
        # Sort dates in descending order (newest first)
        dates = sorted(list(dates_set), reverse=True)
        
        print(f"Found dates: {dates}")
        
        return jsonify({'success': True, 'dates': dates})
        
    except Exception as e:
        print(f"Error getting history dates: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@exams_bp.route('/api/marks/history/snapshot', methods=['POST'])
@login_required
def create_history_snapshot():
    """Manually create a history snapshot of current marks"""
    user = session.get('user')
    institute = get_institute(user['id'])
    
    if not institute:
        return jsonify({'success': False, 'message': 'Institute not found'}), 400
    
    try:
        data = request.get_json()
        exam_id = data.get('exam_id')
        class_id = data.get('class_id')
        
        if not exam_id or not class_id:
            return jsonify({'success': False, 'message': 'Exam ID and Class ID required'}), 400
        
        # Get all current marks
        marks_response = supabase.table('exam_marks')\
            .select('*')\
            .eq('exam_id', exam_id)\
            .eq('class_id', class_id)\
            .eq('institute_id', institute['id'])\
            .execute()
        
        if not marks_response.data:
            return jsonify({'success': False, 'message': 'No marks to snapshot'}), 400
        
        snapshot_count = 0
        for mark in marks_response.data:
            # Check if already have a snapshot for today
            existing = supabase.table('exam_marks_history')\
                .select('id')\
                .eq('exam_id', exam_id)\
                .eq('class_id', class_id)\
                .eq('student_id', mark['student_id'])\
                .eq('subject_id', mark['subject_id'])\
                .eq('record_date', datetime.now().date().isoformat())\
                .execute()
            
            if not existing.data:
                history_data = {
                    'id': str(uuid.uuid4()),
                    'institute_id': institute['id'],
                    'exam_id': exam_id,
                    'class_id': class_id,
                    'student_id': mark['student_id'],
                    'subject_id': mark['subject_id'],
                    'obtained_marks': mark['obtained_marks'],
                    'record_date': datetime.now().date().isoformat(),
                    'created_at': datetime.now().isoformat()
                }
                supabase.table('exam_marks_history').insert(history_data).execute()
                snapshot_count += 1
        
        return jsonify({
            'success': True,
            'message': f'Created snapshot for {snapshot_count} mark(s)',
            'snapshot_count': snapshot_count
        })
        
    except Exception as e:
        print(f"Error creating history snapshot: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@exams_bp.route('/api/marksheet/excel', methods=['POST'])
@login_required
def export_marksheet_excel():
    user = session.get('user')
    institute = get_institute(user['id'])
    
    if not institute:
        return jsonify({'success': False, 'message': 'Institute not found'}), 400
    
    try:
        data = request.get_json()
        exam_id = data.get('exam_id')
        class_id = data.get('class_id')
        academic_year = data.get('academic_year', datetime.now().year)
        version = data.get('version', 'current')
        history_date = data.get('history_date')
        
        if not exam_id or not class_id:
            return jsonify({'success': False, 'message': 'Exam ID and Class ID required'}), 400
        
        # Get exam details
        exam_response = supabase.table('exams')\
            .select('*')\
            .eq('id', exam_id)\
            .execute()
        exam = exam_response.data[0] if exam_response.data else None
        
        # Get class details
        class_response = supabase.table('classes')\
            .select('*')\
            .eq('id', class_id)\
            .execute()
        class_info = class_response.data[0] if class_response.data else None
        
        # Get students
        enrollments_response = supabase.table('class_enrollments')\
            .select('student_id')\
            .eq('class_id', class_id)\
            .eq('academic_year', int(academic_year))\
            .execute()
        
        student_ids = [e['student_id'] for e in enrollments_response.data] if enrollments_response.data else []
        
        all_students = []
        if student_ids:
            students_response = supabase.table('students')\
                .select('id, name, student_id')\
                .eq('institute_id', institute['id'])\
                .in_('id', student_ids)\
                .order('name')\
                .execute()
            all_students = students_response.data if students_response.data else []
        
        # Get subjects
        subjects_response = supabase.table('class_subjects')\
            .select('*, subjects(name)')\
            .eq('class_id', class_id)\
            .eq('institute_id', institute['id'])\
            .order('created_at')\
            .execute()
        subjects = subjects_response.data if subjects_response.data else []
        
        # Get marks based on version
        if version == 'historical' and history_date:
            marks_response = supabase.table('exam_marks_history')\
                .select('*')\
                .eq('exam_id', exam_id)\
                .eq('class_id', class_id)\
                .eq('institute_id', institute['id'])\
                .eq('record_date', history_date)\
                .execute()
        else:
            if student_ids:
                marks_response = supabase.table('exam_marks')\
                    .select('*')\
                    .eq('exam_id', exam_id)\
                    .eq('class_id', class_id)\
                    .eq('institute_id', institute['id'])\
                    .in_('student_id', student_ids)\
                    .execute()
            else:
                marks_response = supabase.table('exam_marks')\
                    .select('*')\
                    .eq('exam_id', exam_id)\
                    .eq('class_id', class_id)\
                    .eq('institute_id', institute['id'])\
                    .execute()
        
        # Group marks by student
        students_marks = {}
        for mark in marks_response.data if marks_response.data else []:
            student_id = mark['student_id']
            if student_id not in students_marks:
                students_marks[student_id] = {'marks': {}}
            students_marks[student_id]['marks'][mark['subject_id']] = mark['obtained_marks']
        
        # Calculate statistics
        percentages = []
        highest_score = 0
        lowest_score = 100
        
        # Create DataFrame
        data_rows = []
        for idx, student in enumerate(all_students, 1):
            row = {'S/N': idx, 'Student Name': student['name'], 'Student ID': student['student_id']}
            total_obtained = 0
            total_max = 0
            
            for subject in subjects:
                subject_name = subject['subjects']['name'] if subject.get('subjects') else 'N/A'
                max_marks = subject['marks']
                total_max += max_marks
                
                if student['id'] in students_marks:
                    obtained = students_marks[student['id']]['marks'].get(subject['subject_id'], '-')
                    if obtained != '-':
                        total_obtained += obtained
                    row[subject_name] = obtained if obtained != '-' else '-'
                else:
                    row[subject_name] = '-'
            
            percentage = round((total_obtained / total_max * 100), 1) if total_max > 0 else 0
            percentages.append(percentage)
            if percentage > highest_score:
                highest_score = percentage
            if percentage < lowest_score:
                lowest_score = percentage
            
            row['Total'] = f"{total_obtained}/{total_max}"
            row['Percentage'] = f"{percentage}%"
            data_rows.append(row)
        
        if not data_rows:
            return jsonify({'success': False, 'message': 'No data to export'}), 404
        
        df = pd.DataFrame(data_rows)
        class_average = sum(percentages) / len(percentages) if percentages else 0
        
        # Create Excel file
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Marksheet', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Marksheet']
            
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_alignment_center = Alignment(horizontal='center', vertical='center')
            cell_alignment_left = Alignment(horizontal='left', vertical='center')
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Style headers
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                column_letter = get_column_letter(col)
                max_length = max(df[df.columns[col-1]].astype(str).map(len).max(), len(df.columns[col-1])) + 2
                worksheet.column_dimensions[column_letter].width = min(max_length, 25)
            
            # Style data cells
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = cell_alignment_center if col != 2 else cell_alignment_left
                    if row % 2 == 0:
                        cell.fill = PatternFill(start_color='f8fafc', end_color='f8fafc', fill_type='solid')
            
            # Add title row
            worksheet.insert_rows(1)
            worksheet.row_dimensions[1].height = 30
            institute_cell = worksheet.cell(row=1, column=1, value=institute.get('institute_name', 'SCHOOL NAME'))
            institute_cell.font = Font(name='Calibri', size=16, bold=True, color='1e3a5f')
            institute_cell.alignment = Alignment(horizontal='center', vertical='center')
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
            
            # Add info row
            worksheet.insert_rows(2)
            version_text = f"{version.upper()} MARKS" if version == 'current' else f"HISTORICAL MARKS - {history_date}"
            info_cell = worksheet.cell(row=2, column=1, value=f"EXAMINATION: {exam['exam_name'] if exam else 'N/A'} | CLASS: {class_info['name'] if class_info else 'N/A'} | ACADEMIC YEAR: {academic_year} | {version_text} | DATE: {datetime.now().strftime('%d %B, %Y')}")
            info_cell.font = Font(name='Calibri', size=10, italic=True, color='4b5563')
            info_cell.alignment = Alignment(horizontal='center', vertical='center')
            worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(df.columns))
            
            # Adjust header row
            for col in range(1, len(df.columns) + 1):
                header_cell = worksheet.cell(row=3, column=col)
                header_cell.font = header_font
                header_cell.fill = header_fill
                header_cell.alignment = header_alignment
                header_cell.border = thin_border
            
            # Add summary sheet
            summary_data = {
                'Metric': ['Version', 'Academic Year', 'Total Students', 'Total Subjects', 'Class Average', 'Highest Score', 'Lowest Score'],
                'Value': [
                    version_text,
                    academic_year,
                    len(all_students),
                    len(subjects),
                    f"{class_average:.1f}%",
                    f"{highest_score:.1f}%",
                    f"{lowest_score:.1f}%"
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            summary_ws = writer.sheets['Summary']
            for col in range(1, 3):
                summary_ws.column_dimensions[get_column_letter(col)].width = 25
                header_cell = summary_ws.cell(row=1, column=col)
                header_cell.font = header_font
                header_cell.fill = header_fill
                header_cell.alignment = header_alignment
        
        output.seek(0)
        
        filename = f"MARKSHEET_{exam['exam_name']}_{class_info['name']}_{academic_year}_{version}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except Exception as e:
        print(f"Error generating marksheet Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500
    
    
@exams_bp.route('/api/marks/historical', methods=['GET'])
@login_required
def get_historical_marks():
    """Get historical marks for an exam and class on a specific date"""
    user = session.get('user')
    institute = get_institute(user['id'])
    
    if not institute:
        return jsonify({'success': False, 'message': 'Institute not found'}), 400
    
    try:
        exam_id = request.args.get('exam_id')
        class_id = request.args.get('class_id')
        academic_year = request.args.get('academic_year', str(datetime.now().year))
        history_date = request.args.get('history_date')
        
        print(f"Getting historical marks - exam: {exam_id}, class: {class_id}, date: {history_date}")
        
        if not exam_id or not class_id:
            return jsonify({'success': False, 'message': 'Exam ID and Class ID required'}), 400
        
        if not history_date:
            return jsonify({'success': False, 'message': 'History date required'}), 400
        
        # Get historical marks for the specific date
        marks_response = supabase.table('exam_marks_history')\
            .select('*')\
            .eq('exam_id', exam_id)\
            .eq('class_id', class_id)\
            .eq('institute_id', institute['id'])\
            .eq('record_date', history_date)\
            .execute()
        
        print(f"Found {len(marks_response.data) if marks_response.data else 0} historical marks")
        
        if not marks_response.data:
            return jsonify({'success': True, 'students': [], 'subjects': [], 'history_date': history_date})
        
        # Get unique student IDs from the history records
        student_ids = list(set([mark['student_id'] for mark in marks_response.data]))
        
        print(f"Student IDs from history: {student_ids}")
        
        # Get student details
        students_response = supabase.table('students')\
            .select('id, name, student_id')\
            .eq('institute_id', institute['id'])\
            .in_('id', student_ids)\
            .order('name')\
            .execute()
        
        students = students_response.data if students_response.data else []
        
        print(f"Found {len(students)} students")
        
        # Get subjects for the class
        subjects_response = supabase.table('class_subjects')\
            .select('*, subjects(name)')\
            .eq('class_id', class_id)\
            .eq('institute_id', institute['id'])\
            .execute()
        
        subjects = subjects_response.data if subjects_response.data else []
        
        print(f"Found {len(subjects)} subjects")
        
        # Group marks by student
        historical_marks = {}
        for mark in marks_response.data:
            student_id = mark['student_id']
            if student_id not in historical_marks:
                historical_marks[student_id] = {}
            historical_marks[student_id][mark['subject_id']] = mark['obtained_marks']
        
        # Prepare marks data
        marks_data = []
        for student in students:
            student_marks = {
                'student_id': student['id'],
                'student_name': student['name'],
                'student_number': student['student_id'],
                'subjects': []
            }
            total_obtained = 0
            total_max = 0
            
            for subject in subjects:
                subject_id = subject['subject_id']
                subject_name = subject['subjects']['name'] if subject.get('subjects') else 'N/A'
                max_marks = subject['marks']
                
                if student['id'] in historical_marks and subject_id in historical_marks[student['id']]:
                    obtained = historical_marks[student['id']][subject_id]
                else:
                    obtained = None
                
                student_marks['subjects'].append({
                    'subject_id': subject_id,
                    'subject_name': subject_name,
                    'max_marks': max_marks,
                    'obtained': obtained
                })
                
                if obtained:
                    total_obtained += obtained
                    total_max += max_marks
            
            student_marks['total_obtained'] = total_obtained
            student_marks['total_max'] = total_max
            student_marks['percentage'] = round((total_obtained / total_max * 100), 1) if total_max > 0 else 0
            marks_data.append(student_marks)
        
        return jsonify({
            'success': True,
            'students': marks_data,
            'subjects': [{'id': s['subject_id'], 'name': s['subjects']['name'], 'max_marks': s['marks']} for s in subjects],
            'history_date': history_date
        })
        
    except Exception as e:
        print(f"Error getting historical marks: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500
    
    
# Add these new endpoints to exams.py

@exams_bp.route('/api/marksheet/generate-id', methods=['POST'])
@login_required
def generate_marksheet_id():
    """Generate a unique marksheet ID for the current marks entry session"""
    user = session.get('user')
    institute = get_institute(user['id'])
    
    if not institute:
        return jsonify({'success': False, 'message': 'Institute not found'}), 400
    
    try:
        data = request.get_json()
        exam_id = data.get('exam_id')
        class_id = data.get('class_id')
        academic_year = data.get('academic_year', datetime.now().year)
        
        if not exam_id or not class_id:
            return jsonify({'success': False, 'message': 'Exam ID and Class ID required'}), 400
        
        # Get exam details to get total marks
        exam_response = supabase.table('exams')\
            .select('total_marks, exam_name')\
            .eq('id', exam_id)\
            .execute()
        
        if not exam_response.data:
            return jsonify({'success': False, 'message': 'Exam not found'}), 404
        
        exam = exam_response.data[0]
        
        # Generate marksheet ID
        marksheet_id = str(uuid.uuid4())
        marksheet_number = f"MS-{datetime.now().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
        
        # Create marksheet record
        marksheet_data = {
            'id': marksheet_id,
            'institute_id': institute['id'],
            'exam_id': exam_id,
            'class_id': class_id,
            'academic_year': academic_year,
            'marksheet_number': marksheet_number,
            'generated_at': datetime.now().isoformat(),
            'created_at': datetime.now().isoformat()
        }
        
        supabase.table('exam_marksheets').insert(marksheet_data).execute()
        
        return jsonify({
            'success': True,
            'marksheet_id': marksheet_id,
            'marksheet_number': marksheet_number,
            'exam_total_marks': exam['total_marks']
        })
        
    except Exception as e:
        print(f"Error generating marksheet ID: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@exams_bp.route('/api/marks/save-with-marksheet', methods=['POST'])
@login_required
def save_marks_with_marksheet():
    """Save exam marks with marksheet ID for historical tracking"""
    user = session.get('user')
    institute = get_institute(user['id'])
    
    if not institute:
        return jsonify({'success': False, 'message': 'Institute not found'}), 400
    
    try:
        data = request.get_json()
        exam_id = data.get('exam_id')
        class_id = data.get('class_id')
        academic_year = data.get('academic_year', datetime.now().year)
        marksheet_id = data.get('marksheet_id')
        marks_data = data.get('marks', [])
        
        if not exam_id or not class_id:
            return jsonify({'success': False, 'message': 'Exam ID and Class ID required'}), 400
        
        # Get exam total marks
        exam_response = supabase.table('exams')\
            .select('total_marks')\
            .eq('id', exam_id)\
            .execute()
        
        exam_total_marks = exam_response.data[0]['total_marks'] if exam_response.data else 0
        
        # If no marksheet_id provided, create one
        if not marksheet_id:
            marksheet_id = str(uuid.uuid4())
            marksheet_number = f"MS-{datetime.now().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
            
            marksheet_data = {
                'id': marksheet_id,
                'institute_id': institute['id'],
                'exam_id': exam_id,
                'class_id': class_id,
                'academic_year': academic_year,
                'marksheet_number': marksheet_number,
                'generated_at': datetime.now().isoformat(),
                'created_at': datetime.now().isoformat()
            }
            supabase.table('exam_marksheets').insert(marksheet_data).execute()
        
        saved_count = 0
        errors = []
        
        for mark_entry in marks_data:
            try:
                student_id = mark_entry.get('student_id')
                subject_id = mark_entry.get('subject_id')
                obtained_marks = mark_entry.get('obtained_marks')
                
                if obtained_marks is None or obtained_marks == '':
                    continue
                
                obtained_marks = float(obtained_marks)
                
                # Check if record exists for this marksheet
                existing = supabase.table('exam_marks')\
                    .select('id')\
                    .eq('exam_id', exam_id)\
                    .eq('class_id', class_id)\
                    .eq('student_id', student_id)\
                    .eq('subject_id', subject_id)\
                    .eq('marksheet_id', marksheet_id)\
                    .eq('institute_id', institute['id'])\
                    .execute()
                
                if existing.data:
                    # Update existing
                    supabase.table('exam_marks')\
                        .update({
                            'obtained_marks': obtained_marks,
                            'updated_at': datetime.now().isoformat()
                        })\
                        .eq('id', existing.data[0]['id'])\
                        .execute()
                else:
                    # Insert new with marksheet_id
                    mark_id = str(uuid.uuid4())
                    mark_data = {
                        'id': mark_id,
                        'institute_id': institute['id'],
                        'exam_id': exam_id,
                        'class_id': class_id,
                        'student_id': student_id,
                        'subject_id': subject_id,
                        'obtained_marks': obtained_marks,
                        'marksheet_id': marksheet_id,
                        'exam_total_marks': exam_total_marks,
                        'created_at': datetime.now().isoformat(),
                        'updated_at': datetime.now().isoformat()
                    }
                    supabase.table('exam_marks').insert(mark_data).execute()
                
                saved_count += 1
                
            except Exception as e:
                errors.append(f"Error saving marks for student {mark_entry.get('student_id')}: {str(e)}")
        
        return jsonify({
            'success': True,
            'message': f'Saved {saved_count} mark(s) successfully',
            'marksheet_id': marksheet_id,
            'errors': errors if errors else None
        })
        
    except Exception as e:
        print(f"Error saving marks: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@exams_bp.route('/api/marks/history/by-student', methods=['GET'])
@login_required
def get_student_marks_history():
    """Get historical marks for a student across different marksheets"""
    user = session.get('user')
    institute = get_institute(user['id'])
    
    if not institute:
        return jsonify({'success': False, 'message': 'Institute not found'}), 400
    
    try:
        student_id = request.args.get('student_id')
        
        if not student_id:
            return jsonify({'success': False, 'message': 'Student ID required'}), 400
        
        # Get all marks for this student with marksheet info
        response = supabase.table('exam_marks')\
            .select('*, exams(exam_name, exam_date, total_marks), exam_marksheets(marksheet_number, generated_at)')\
            .eq('student_id', student_id)\
            .eq('institute_id', institute['id'])\
            .order('created_at', desc=True)\
            .execute()
        
        marks = response.data if response.data else []
        
        # Group by marksheet
        marksheets = {}
        for mark in marks:
            marksheet_id = mark.get('marksheet_id')
            if marksheet_id not in marksheets:
                marksheets[marksheet_id] = {
                    'marksheet_id': marksheet_id,
                    'marksheet_number': mark.get('exam_marksheets', {}).get('marksheet_number') if mark.get('exam_marksheets') else 'N/A',
                    'generated_at': mark.get('exam_marksheets', {}).get('generated_at') if mark.get('exam_marksheets') else mark.get('created_at'),
                    'exam_name': mark.get('exams', {}).get('exam_name') if mark.get('exams') else 'N/A',
                    'exam_date': mark.get('exams', {}).get('exam_date') if mark.get('exams') else 'N/A',
                    'exam_total_marks': mark.get('exam_total_marks', 0),
                    'subjects': []
                }
            
            # Get subject name
            subject_response = supabase.table('subjects')\
                .select('name')\
                .eq('id', mark['subject_id'])\
                .execute()
            
            subject_name = subject_response.data[0]['name'] if subject_response.data else 'N/A'
            
            marksheets[marksheet_id]['subjects'].append({
                'subject_name': subject_name,
                'obtained_marks': mark['obtained_marks']
            })
        
        return jsonify({
            'success': True,
            'marksheets': list(marksheets.values())
        })
        
    except Exception as e:
        print(f"Error getting student marks history: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    
@exams_bp.route('/api/marks/by-marksheet', methods=['GET'])
@login_required
def get_marks_by_marksheet():
    """Get marks for a specific marksheet - works regardless of current student class"""
    user = session.get('user')
    institute = get_institute(user['id'])
    
    if not institute:
        return jsonify({'success': False, 'message': 'Institute not found'}), 400
    
    try:
        marksheet_id = request.args.get('marksheet_id')
        
        if not marksheet_id:
            return jsonify({'success': False, 'message': 'Marksheet ID required'}), 400
        
        # Get marksheet details
        marksheet_response = supabase.table('exam_marksheets')\
            .select('*, exams(exam_name, total_marks, exam_date), classes(name)')\
            .eq('id', marksheet_id)\
            .eq('institute_id', institute['id'])\
            .execute()
        
        if not marksheet_response.data:
            return jsonify({'success': False, 'message': 'Marksheet not found'}), 404
        
        marksheet = marksheet_response.data[0]
        exam = marksheet.get('exams', {})
        exam_total_marks = exam.get('total_marks', 0)
        
        # Get all marks for this marksheet
        marks_response = supabase.table('exam_marks')\
            .select('*, students(name, student_id)')\
            .eq('marksheet_id', marksheet_id)\
            .eq('institute_id', institute['id'])\
            .execute()
        
        marks = marks_response.data if marks_response.data else []
        
        # Get subjects for the class
        subjects_response = supabase.table('class_subjects')\
            .select('*, subjects(name)')\
            .eq('class_id', marksheet['class_id'])\
            .eq('institute_id', institute['id'])\
            .execute()
        
        subjects = subjects_response.data if subjects_response.data else []
        
        # Group marks by student
        students_marks = {}
        for mark in marks:
            student_id = mark['student_id']
            if student_id not in students_marks:
                students_marks[student_id] = {
                    'student_id': student_id,
                    'student_name': mark['students']['name'] if mark.get('students') else 'N/A',
                    'student_number': mark['students']['student_id'] if mark.get('students') else 'N/A',
                    'marks': {}
                }
            students_marks[student_id]['marks'][mark['subject_id']] = mark['obtained_marks']
        
        # Prepare marks data
        marks_data = []
        for student_id, student_data in students_marks.items():
            student_marks = {
                'student_id': student_data['student_id'],
                'student_name': student_data['student_name'],
                'student_number': student_data['student_number'],
                'subjects': []
            }
            
            total_obtained = 0
            
            for subject in subjects:
                subject_id = subject['subject_id']
                subject_name = subject['subjects']['name'] if subject.get('subjects') else 'N/A'
                max_marks = subject['marks']
                
                obtained = student_data['marks'].get(subject_id)
                
                student_marks['subjects'].append({
                    'subject_id': subject_id,
                    'subject_name': subject_name,
                    'max_marks': max_marks,
                    'obtained': obtained
                })
                
                if obtained:
                    total_obtained += obtained
            
            percentage = round((total_obtained / exam_total_marks * 100), 1) if exam_total_marks > 0 else 0
            
            student_marks['total_obtained'] = total_obtained
            student_marks['exam_total_marks'] = exam_total_marks
            student_marks['percentage'] = percentage
            marks_data.append(student_marks)
        
        return jsonify({
            'success': True,
            'marksheet': {
                'id': marksheet['id'],
                'marksheet_number': marksheet['marksheet_number'],
                'exam_name': exam.get('exam_name', 'N/A'),
                'exam_date': exam.get('exam_date', 'N/A'),
                'class_name': marksheet.get('classes', {}).get('name', 'N/A'),
                'academic_year': marksheet['academic_year'],
                'generated_at': marksheet['generated_at']
            },
            'students': marks_data,
            'subjects': [{'id': s['subject_id'], 'name': s['subjects']['name'], 'max_marks': s['marks']} for s in subjects],
            'exam_total_marks': exam_total_marks
        })
        
    except Exception as e:
        print(f"Error getting marks by marksheet: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@exams_bp.route('/api/marksheets/list', methods=['GET'])
@login_required
def list_marksheets():
    """List all marksheets for an institute"""
    user = session.get('user')
    institute = get_institute(user['id'])
    
    if not institute:
        return jsonify({'success': False, 'message': 'Institute not found'}), 400
    
    try:
        exam_id = request.args.get('exam_id')
        
        query = supabase.table('exam_marksheets')\
            .select('*, exams(exam_name), classes(name)')\
            .eq('institute_id', institute['id'])
        
        if exam_id:
            query = query.eq('exam_id', exam_id)
        
        response = query.order('generated_at', desc=True).execute()
        
        marksheets = response.data if response.data else []
        
        return jsonify({'success': True, 'marksheets': marksheets})
        
    except Exception as e:
        print(f"Error listing marksheets: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500